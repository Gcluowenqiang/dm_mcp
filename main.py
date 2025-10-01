#!/usr/bin/env python3
"""
达梦数据库MCP服务
专为Cursor设计，提供表结构查询和文档生成功能
支持多种安全模式

Copyright (c) 2025 qyue
Licensed under the MIT License.
See LICENSE file in the project root for full license information.
"""

import asyncio
import json
import sys
import os
from datetime import datetime
from typing import Any, Sequence
import logging

from mcp.server.models import InitializationOptions
from mcp.server import NotificationOptions, Server
from mcp.server.stdio import stdio_server
from mcp.server.fastmcp import FastMCP
from mcp.types import (
    Resource, 
    Tool, 
    TextContent, 
    ImageContent, 
    EmbeddedResource, 
    LoggingLevel
)
from pydantic import AnyUrl

from database import get_db_instance
from document_generator import doc_generator
from config import get_config_instance

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def normalize_data(data_list):
    """标准化数据，将大写字段名转换为小写（保持原字段名作为备份）"""
    normalized = []
    for item in data_list:
        normalized_item = {}
        for key, value in item.items():
            # 保留原字段名
            normalized_item[key] = value
            # 添加小写字段名
            normalized_item[key.lower()] = value
        normalized.append(normalized_item)
    return normalized

def create_error_response(error_msg: str, error_type: str = "error") -> list[TextContent]:
    """创建统一的错误响应"""
    logger.error(f"{error_type}: {error_msg}")
    return [TextContent(
        type="text",
        text=f"❌ {error_type.upper()}: {error_msg}"
    )]

def create_success_response(success_msg: str) -> list[TextContent]:
    """创建统一的成功响应"""
    logger.info(f"Success: {success_msg}")
    return [TextContent(
        type="text",
        text=f"✅ {success_msg}"
    )]

# 创建MCP服务器
server = Server("dm-mcp")

@server.list_tools()
async def handle_list_tools() -> list[Tool]:
    """
    列出可用的工具
    """
    return [
        Tool(
            name="test_connection",
            description="测试达梦数据库连接",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        Tool(
            name="get_security_info",
            description="获取当前安全配置信息",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        Tool(
            name="list_tables",
            description="获取数据库中所有表的列表",
            inputSchema={
                "type": "object",
                "properties": {
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="describe_table",
            description="获取指定表的详细结构信息",
            inputSchema={
                "type": "object",
                "properties": {
                    "table_name": {
                        "type": "string",
                        "description": "表名"
                    },
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    }
                },
                "required": ["table_name"]
            }
        ),
        Tool(
            name="generate_table_doc",
            description="生成表结构设计文档并保存为文件（支持Markdown、JSON、SQL格式）",
            inputSchema={
                "type": "object",
                "properties": {
                    "table_name": {
                        "type": "string",
                        "description": "表名"
                    },
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    },
                    "format": {
                        "type": "string",
                        "description": "文档格式: markdown, json, sql",
                        "enum": ["markdown", "json", "sql"],
                        "default": "markdown"
                    }
                },
                "required": ["table_name"]
            }
        ),
        Tool(
            name="generate_database_overview",
            description="生成数据库概览文档并保存为Markdown文件",
            inputSchema={
                "type": "object",
                "properties": {
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="execute_query",
            description="执行SQL语句（根据安全模式限制操作类型）",
            inputSchema={
                "type": "object",
                "properties": {
                    "sql": {
                        "type": "string",
                        "description": "SQL语句"
                    }
                },
                "required": ["sql"]
            }
        ),
        Tool(
            name="list_schemas",
            description="获取用户有权限访问的所有数据库模式",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        Tool(
            name="generate_relationship_doc",
            description="生成数据库表关系图文档（支持Mermaid格式）",
            inputSchema={
                "type": "object",
                "properties": {
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    }
                },
                "required": []
            }
        ),
        Tool(
            name="batch_generate_table_docs",
            description="批量生成多个表的文档",
            inputSchema={
                "type": "object",
                "properties": {
                    "table_names": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "表名列表"
                    },
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    },
                    "format": {
                        "type": "string",
                        "description": "文档格式: markdown, json, sql",
                        "enum": ["markdown", "json", "sql"],
                        "default": "markdown"
                    }
                },
                "required": ["table_names"]
            }
        ),
        Tool(
            name="export_to_excel",
            description="导出表结构或数据为Excel格式",
            inputSchema={
                "type": "object",
                "properties": {
                    "table_name": {
                        "type": "string",
                        "description": "表名"
                    },
                    "schema": {
                        "type": "string",
                        "description": "数据库模式名称",
                        "default": "SYSDBA"
                    },
                    "export_type": {
                        "type": "string",
                        "description": "导出类型: structure, data, both",
                        "enum": ["structure", "data", "both"],
                        "default": "structure"
                    },
                    "data_limit": {
                        "type": "number",
                        "description": "数据导出行数限制",
                        "default": 1000
                    },
                    "fast_mode": {
                        "type": "boolean",
                        "description": "快速模式（禁用样式，提高导出速度）",
                        "default": True
                    }
                },
                "required": ["table_name"]
            }
        ),
        Tool(
            name="get_cache_info",
            description="获取查询缓存统计信息",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        ),
        Tool(
            name="clear_cache",
            description="清空查询缓存",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        )
    ]

@server.call_tool()
async def handle_call_tool(name: str, arguments: dict[str, Any] | None) -> list[TextContent | ImageContent | EmbeddedResource]:
    """
    处理工具调用
    """
    try:
        # 获取数据库实例
        db = get_db_instance()
        
        if name == "test_connection":
            result = db.test_connection()
            return [TextContent(
                type="text",
                text=f"达梦数据库连接测试: {'成功' if result else '失败'}"
            )]
        
        elif name == "get_security_info":
            security_info = db.get_security_info()
            info_text = "当前安全配置信息:\n\n"
            info_text += f"安全模式: {security_info['security_mode']}\n"
            info_text += f"只读模式: {'是' if security_info['readonly_mode'] else '否'}\n"
            info_text += f"允许写入操作: {'是' if security_info['write_allowed'] else '否'}\n"
            info_text += f"允许危险操作: {'是' if security_info['dangerous_operations_allowed'] else '否'}\n"
            info_text += f"允许访问的模式: {', '.join(security_info['allowed_schemas'])}\n"
            info_text += f"最大返回行数: {security_info['max_result_rows']}\n"
            info_text += f"查询日志: {'启用' if security_info['query_log_enabled'] else '禁用'}\n"
            
            return [TextContent(type="text", text=info_text)]
        
        elif name == "list_tables":
            schema = arguments.get("schema", "SYSDBA") if arguments else "SYSDBA"
            tables = db.get_all_tables(schema)
            
            if not tables:
                return [TextContent(
                    type="text",
                    text=f"在模式 '{schema}' 中没有找到任何表"
                )]
            
            # 达梦数据库字段名可能是大写，尝试两种格式
            table_list = "\n".join([f"- {table.get('tablename') or table.get('TABLENAME', 'Unknown')}" for table in tables])
            return [TextContent(
                type="text",
                text=f"模式 '{schema}' 中的表列表:\n{table_list}\n\n总计: {len(tables)} 个表"
            )]
        
        elif name == "describe_table":
            if not arguments or "table_name" not in arguments:
                return create_error_response("缺少必需的参数 'table_name'", "参数错误")
            
            table_name = arguments["table_name"]
            schema = arguments.get("schema", "SYSDBA")
            
            # 获取表结构信息
            structure = db.get_table_structure(table_name, schema)
            indexes = db.get_table_indexes(table_name, schema)
            constraints = db.get_table_constraints(table_name, schema)
            table_comment = db.get_table_comment(table_name, schema)
            
            if not structure:
                return [TextContent(
                    type="text",
                    text=f"表 '{table_name}' 在模式 '{schema}' 中不存在"
                )]
            
            # 格式化输出
            result = f"表 '{table_name}' 结构信息:\n\n"
            if table_comment:
                result += f"表注释: {table_comment}\n\n"
            result += "字段列表:\n"
            for col in structure:
                # 达梦数据库字段名可能是大写，尝试两种格式
                column_name = col.get('column_name') or col.get('COLUMN_NAME', 'Unknown')
                data_type = col.get('data_type') or col.get('DATA_TYPE', 'Unknown')
                is_nullable = col.get('is_nullable') or col.get('IS_NULLABLE', 'YES')
                is_primary_key = col.get('is_primary_key') or col.get('IS_PRIMARY_KEY', 'NO')
                column_comment = col.get('column_comment') or col.get('COLUMN_COMMENT', '')
                
                result += f"- {column_name} ({data_type}) "
                if is_nullable == 'NO':
                    result += "NOT NULL "
                if is_primary_key == 'YES':
                    result += "[主键] "
                if column_comment:
                    result += f"-- {column_comment}"
                result += "\n"
            
            if indexes:
                result += f"\n索引 ({len(indexes)} 个):\n"
                for idx in indexes:
                    # 达梦数据库字段名可能是大写，尝试两种格式
                    indexname = idx.get('indexname') or idx.get('INDEXNAME', 'Unknown')
                    is_unique = idx.get('is_unique') or idx.get('IS_UNIQUE', 'NO')
                    result += f"- {indexname} {'[唯一]' if is_unique == 'YES' else ''}\n"
            
            if constraints:
                result += f"\n约束 ({len(constraints)} 个):\n"
                for constraint in constraints:
                    # 达梦数据库字段名可能是大写，尝试两种格式
                    constraint_name = constraint.get('constraint_name') or constraint.get('CONSTRAINT_NAME', 'Unknown')
                    constraint_type = constraint.get('constraint_type') or constraint.get('CONSTRAINT_TYPE', 'Unknown')
                    result += f"- {constraint_name} ({constraint_type})\n"
            
            return [TextContent(type="text", text=result)]
        
        elif name == "generate_table_doc":
            if not arguments or "table_name" not in arguments:
                return create_error_response("缺少必需的参数 'table_name'", "参数错误")
            
            try:
                table_name = arguments["table_name"]
                schema = arguments.get("schema", "SYSDBA")
                format_type = arguments.get("format", "markdown")
                
                # 获取表信息
                structure = db.get_table_structure(table_name, schema)
                indexes = db.get_table_indexes(table_name, schema)
                constraints = db.get_table_constraints(table_name, schema)
                table_comment = db.get_table_comment(table_name, schema)
                
                if not structure:
                    return [TextContent(
                        type="text",
                        text=f"表 '{table_name}' 在模式 '{schema}' 中不存在"
                    )]
                
                # 预处理数据，确保字段名兼容性
                structure = normalize_data(structure)
                indexes = normalize_data(indexes)
                constraints = normalize_data(constraints)
                
                # 生成文档
                if format_type == "markdown":
                    doc = doc_generator.generate_table_structure_doc(table_name, structure, indexes, constraints, schema, table_comment)
                    file_ext = ".md"
                elif format_type == "json":
                    doc = doc_generator.generate_json_structure(table_name, structure, indexes, constraints, schema, table_comment)
                    file_ext = ".json"
                elif format_type == "sql":
                    doc = doc_generator.generate_sql_create_statement(table_name, structure, table_comment)
                    file_ext = ".sql"
                else:
                    return [TextContent(
                        type="text",
                        text=f"不支持的文档格式: {format_type}"
                    )]
                
                # 确保在MCP服务目录下创建docs目录
                service_dir = os.path.dirname(os.path.abspath(__file__))  # 获取MCP服务所在目录
                docs_dir = os.path.join(service_dir, "docs")
                os.makedirs(docs_dir, exist_ok=True)
                
                # 生成文件名
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{schema}_{table_name}_{timestamp}{file_ext}"
                file_path = os.path.join(docs_dir, filename)
                
                # 保存文档到文件
                try:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(doc)
                    
                    # 返回成功信息和文档预览
                    # 显示MCP服务目录的相对路径
                    relative_path = os.path.relpath(file_path, service_dir)
                    result_text = f"✅ 文档生成成功!\n\n"
                    result_text += f"📁 保存路径: {relative_path}\n"
                    result_text += f"📂 MCP服务目录: {service_dir}\n"
                    result_text += f"📊 表名: {schema}.{table_name}\n"
                    result_text += f"📝 格式: {format_type}\n"
                    result_text += f"⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    result_text += "📄 文档内容预览:\n"
                    result_text += "=" * 50 + "\n"
                    
                    # 限制预览长度
                    preview = doc[:1000] + "..." if len(doc) > 1000 else doc
                    result_text += preview
                    
                    return [TextContent(type="text", text=result_text)]
                    
                except Exception as file_error:
                    # 如果文件保存失败，仍然返回文档内容
                    error_msg = f"⚠️ 文件保存失败: {str(file_error)}\n\n"
                    error_msg += "📄 生成的文档内容:\n"
                    error_msg += "=" * 50 + "\n"
                    error_msg += doc
                    return [TextContent(type="text", text=error_msg)]
            
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"生成文档时发生错误: {str(e)}"
                )]
        
        elif name == "generate_database_overview":
            try:
                schema = arguments.get("schema", "SYSDBA") if arguments else "SYSDBA"
                tables = db.get_all_tables(schema)
                
                # 预处理数据，确保字段名兼容性
                tables = normalize_data(tables)
                
                # 生成数据库概览文档
                doc = doc_generator.generate_database_overview_doc(tables, schema)
                
                # 确保在MCP服务目录下创建docs目录
                service_dir = os.path.dirname(os.path.abspath(__file__))  # 获取MCP服务所在目录
                docs_dir = os.path.join(service_dir, "docs")
                os.makedirs(docs_dir, exist_ok=True)
                
                # 生成文件名
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{schema}_数据库概览_{timestamp}.md"
                file_path = os.path.join(docs_dir, filename)
                
                # 保存文档到文件
                try:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(doc)
                    
                    # 返回成功信息和文档预览
                    # 显示MCP服务目录的相对路径
                    relative_path = os.path.relpath(file_path, service_dir)
                    result_text = f"✅ 数据库概览文档生成成功!\n\n"
                    result_text += f"📁 保存路径: {relative_path}\n"
                    result_text += f"📂 MCP服务目录: {service_dir}\n"
                    result_text += f"🗂️ 模式: {schema}\n"
                    result_text += f"📋 表数量: {len(tables)} 个\n"
                    result_text += f"⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    result_text += "📄 文档内容预览:\n"
                    result_text += "=" * 50 + "\n"
                    
                    # 限制预览长度
                    preview = doc[:1000] + "..." if len(doc) > 1000 else doc
                    result_text += preview
                    
                    return [TextContent(type="text", text=result_text)]
                    
                except Exception as file_error:
                    # 如果文件保存失败，仍然返回文档内容
                    error_msg = f"⚠️ 文件保存失败: {str(file_error)}\n\n"
                    error_msg += "📄 生成的文档内容:\n"
                    error_msg += "=" * 50 + "\n"
                    error_msg += doc
                    return [TextContent(type="text", text=error_msg)]
                    
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"生成数据库概览文档时发生错误: {str(e)}"
                )]
        
        elif name == "execute_query":
            if not arguments or "sql" not in arguments:
                return create_error_response("缺少必需的参数 'sql'", "参数错误")
            
            sql = arguments["sql"]
            
            try:
                results = db.execute_query(sql)
                
                if not results:
                    return [TextContent(
                        type="text",
                        text="语句执行成功，但没有返回结果"
                    )]
                
                # 格式化结果
                if sql.upper().strip().startswith(('SELECT', 'WITH', 'SHOW', 'DESCRIBE', 'EXPLAIN')):
                    result_text = f"查询结果 ({len(results)} 条记录):\n\n"
                    
                    if len(results) <= 100:  # 限制显示条数
                        result_text += json.dumps(results, ensure_ascii=False, indent=2)
                    else:
                        result_text += f"结果集过大，仅显示前100条:\n"
                        result_text += json.dumps(results[:100], ensure_ascii=False, indent=2)
                        result_text += f"\n\n... (还有 {len(results) - 100} 条记录)"
                else:
                    # 非查询操作的结果
                    result_text = f"操作执行成功:\n\n"
                    result_text += json.dumps(results, ensure_ascii=False, indent=2)
                
                return [TextContent(type="text", text=result_text)]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"SQL执行失败: {str(e)}"
                )]
        
        elif name == "list_schemas":
            try:
                schemas = db.get_available_schemas()
                
                if not schemas:
                    return [TextContent(
                        type="text",
                        text="没有找到可访问的数据库模式"
                    )]
                
                # 达梦数据库字段名可能是大写，尝试两种格式
                schema_list = "\n".join([f"- {schema.get('schemaname') or schema.get('SCHEMANAME', 'Unknown')}" for schema in schemas])
                
                config_info = f"当前schema访问策略: {db._get_allowed_schemas_display()}\n\n"
                result_text = config_info + f"可访问的数据库模式:\n{schema_list}\n\n总计: {len(schemas)} 个模式"
                
                return [TextContent(type="text", text=result_text)]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"获取schema列表失败: {str(e)}"
                )]
        
        elif name == "generate_relationship_doc":
            try:
                schema = arguments.get("schema", "SYSDBA") if arguments else "SYSDBA"
                
                # 获取表列表和关系信息
                tables = db.get_all_tables(schema)
                relationships = db.get_table_relationships(schema)
                
                # 预处理数据
                tables = normalize_data(tables)
                relationships = normalize_data(relationships)
                
                # 生成关系文档
                doc = doc_generator.generate_relationship_doc(tables, relationships, schema)
                
                # 保存文档
                service_dir = os.path.dirname(os.path.abspath(__file__))
                docs_dir = os.path.join(service_dir, "docs")
                os.makedirs(docs_dir, exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{schema}_表关系图_{timestamp}.md"
                file_path = os.path.join(docs_dir, filename)
                
                try:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(doc)
                    
                    relative_path = os.path.relpath(file_path, service_dir)
                    result_text = f"✅ 表关系图文档生成成功!\n\n"
                    result_text += f"📁 保存路径: {relative_path}\n"
                    result_text += f"📂 MCP服务目录: {service_dir}\n"
                    result_text += f"🗂️ 模式: {schema}\n"
                    result_text += f"📋 表数量: {len(tables)} 个\n"
                    result_text += f"🔗 关系数量: {len(relationships)} 个\n"
                    result_text += f"⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    result_text += "📄 文档内容预览:\n"
                    result_text += "=" * 50 + "\n"
                    
                    preview = doc[:1000] + "..." if len(doc) > 1000 else doc
                    result_text += preview
                    
                    return [TextContent(type="text", text=result_text)]
                    
                except Exception as file_error:
                    error_msg = f"⚠️ 文件保存失败: {str(file_error)}\n\n"
                    error_msg += "📄 生成的文档内容:\n"
                    error_msg += "=" * 50 + "\n"
                    error_msg += doc
                    return [TextContent(type="text", text=error_msg)]
                    
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"生成表关系图文档时发生错误: {str(e)}"
                )]
        
        elif name == "batch_generate_table_docs":
            try:
                if not arguments or "table_names" not in arguments:
                    return create_error_response("缺少必需的参数 'table_names'", "参数错误")
                
                table_names = arguments["table_names"]
                schema = arguments.get("schema", "SYSDBA")
                format_type = arguments.get("format", "markdown")
                
                if not isinstance(table_names, list) or not table_names:
                    return create_error_response("table_names 必须是非空列表", "参数错误")
                
                # 批量生成文档
                results = []
                service_dir = os.path.dirname(os.path.abspath(__file__))
                docs_dir = os.path.join(service_dir, "docs")
                os.makedirs(docs_dir, exist_ok=True)
                
                for table_name in table_names:
                    try:
                        # 获取表信息
                        structure = db.get_table_structure(table_name, schema)
                        indexes = db.get_table_indexes(table_name, schema)
                        constraints = db.get_table_constraints(table_name, schema)
                        table_comment = db.get_table_comment(table_name, schema)
                        
                        if not structure:
                            results.append(f"❌ 表 '{table_name}' 不存在")
                            continue
                        
                        # 预处理数据
                        structure = normalize_data(structure)
                        indexes = normalize_data(indexes)
                        constraints = normalize_data(constraints)
                        
                        # 生成文档
                        if format_type == "markdown":
                            doc = doc_generator.generate_table_structure_doc(table_name, structure, indexes, constraints, schema, table_comment)
                            file_ext = ".md"
                        elif format_type == "json":
                            doc = doc_generator.generate_json_structure(table_name, structure, indexes, constraints, schema, table_comment)
                            file_ext = ".json"
                        elif format_type == "sql":
                            doc = doc_generator.generate_sql_create_statement(table_name, structure, table_comment)
                            file_ext = ".sql"
                        else:
                            results.append(f"❌ 表 '{table_name}': 不支持的格式 {format_type}")
                            continue
                        
                        # 保存文档
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"{schema}_{table_name}_{timestamp}{file_ext}"
                        file_path = os.path.join(docs_dir, filename)
                        
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(doc)
                        
                        results.append(f"✅ 表 '{table_name}': 文档已生成")
                        
                    except Exception as table_error:
                        results.append(f"❌ 表 '{table_name}': {str(table_error)}")
                
                # 返回批量处理结果
                result_text = f"✅ 批量文档生成完成!\n\n"
                result_text += f"📂 保存目录: docs/\n"
                result_text += f"🗂️ 模式: {schema}\n"
                result_text += f"📝 格式: {format_type}\n"
                result_text += f"📋 处理表数: {len(table_names)} 个\n"
                result_text += f"⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                result_text += "📊 处理结果:\n"
                result_text += "=" * 50 + "\n"
                result_text += "\n".join(results)
                
                return [TextContent(type="text", text=result_text)]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"批量生成文档时发生错误: {str(e)}"
                )]
        
        elif name == "export_to_excel":
            try:
                if not arguments or "table_name" not in arguments:
                    return create_error_response("缺少必需的参数 'table_name'", "参数错误")
                
                table_name = arguments["table_name"]
                schema = arguments.get("schema", "SYSDBA")
                export_type = arguments.get("export_type", "structure")
                data_limit = arguments.get("data_limit", 1000)
                fast_mode = arguments.get("fast_mode", True)
                
                # 检查是否安装了openpyxl
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    return [TextContent(
                        type="text",
                        text="❌ 导出Excel功能需要安装openpyxl库\n请运行: pip install openpyxl"
                    )]
                
                # 创建Excel工作簿（优化：禁用自动计算）
                wb = openpyxl.Workbook()
                wb.calculation.calcMode = 'manual'  # 禁用自动计算
                
                if export_type in ["structure", "both"]:
                    # 导出表结构
                    structure = db.get_table_structure(table_name, schema)
                    indexes = db.get_table_indexes(table_name, schema)
                    constraints = db.get_table_constraints(table_name, schema)
                    table_comment = db.get_table_comment(table_name, schema)
                    
                    if not structure:
                        return [TextContent(
                            type="text",
                            text=f"❌ 表 '{table_name}' 在模式 '{schema}' 中不存在，无法导出"
                        )]
                    
                    # 创建表结构工作表
                    ws_structure = wb.active
                    ws_structure.title = "表结构"
                    
                    # 设置标题
                    ws_structure['A1'] = f"表结构: {table_name}"
                    if not fast_mode:
                        ws_structure['A1'].font = Font(bold=True, size=14)
                    ws_structure['A2'] = f"模式: {schema}"
                    ws_structure['A3'] = f"表注释: {table_comment or '无'}"
                    ws_structure['A4'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    
                    # 字段信息（优化：批量写入，减少样式设置）
                    headers = ['序号', '字段名', '数据类型', '长度', '精度', '标度', '可空', '默认值', '主键', '注释']
                    
                    # 批量写入标题行
                    for col, header in enumerate(headers, 1):
                        ws_structure.cell(row=6, column=col, value=header)
                    
                    # 批量写入数据行（不设置样式以提高速度）
                    for row_idx, col_data in enumerate(structure, 7):
                        ws_structure.cell(row=row_idx, column=1, value=col_data.get('ordinal_position', ''))
                        ws_structure.cell(row=row_idx, column=2, value=col_data.get('column_name', ''))
                        ws_structure.cell(row=row_idx, column=3, value=col_data.get('data_type', ''))
                        ws_structure.cell(row=row_idx, column=4, value=col_data.get('character_maximum_length', ''))
                        ws_structure.cell(row=row_idx, column=5, value=col_data.get('numeric_precision', ''))
                        ws_structure.cell(row=row_idx, column=6, value=col_data.get('numeric_scale', ''))
                        ws_structure.cell(row=row_idx, column=7, value=col_data.get('is_nullable', ''))
                        ws_structure.cell(row=row_idx, column=8, value=col_data.get('column_default', ''))
                        ws_structure.cell(row=row_idx, column=9, value=col_data.get('is_primary_key', ''))
                        ws_structure.cell(row=row_idx, column=10, value=col_data.get('column_comment', ''))
                
                if export_type in ["data", "both"]:
                    # 导出表数据
                    if export_type == "both":
                        ws_data = wb.create_sheet("表数据")
                    else:
                        ws_data = wb.active
                        ws_data.title = "表数据"
                    
                    # 查询表数据
                    sql = f"SELECT * FROM {schema}.{table_name} LIMIT {data_limit}"
                    data_results = db.execute_query(sql, use_cache=False)
                    
                    if data_results:
                        # 设置标题
                        ws_data['A1'] = f"表数据: {table_name}"
                        if not fast_mode:
                            ws_data['A1'].font = Font(bold=True, size=14)
                        ws_data['A2'] = f"模式: {schema}"
                        ws_data['A3'] = f"数据行数: {len(data_results)} (限制: {data_limit})"
                        ws_data['A4'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        
                        # 写入数据（优化：批量写入，减少样式设置）
                        if data_results:
                            headers = list(data_results[0].keys())
                            # 批量写入标题行
                            for col, header in enumerate(headers, 1):
                                ws_data.cell(row=6, column=col, value=header)
                            
                            # 批量写入数据行（不设置样式以提高速度）
                            for row_idx, data_row in enumerate(data_results, 7):
                                for col_idx, (key, value) in enumerate(data_row.items(), 1):
                                    ws_data.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
                
                # 保存Excel文件
                service_dir = os.path.dirname(os.path.abspath(__file__))
                docs_dir = os.path.join(service_dir, "docs")
                os.makedirs(docs_dir, exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{schema}_{table_name}_导出_{timestamp}.xlsx"
                file_path = os.path.join(docs_dir, filename)
                
                wb.save(file_path)
                
                relative_path = os.path.relpath(file_path, service_dir)
                result_text = f"✅ Excel导出成功!\n\n"
                result_text += f"📁 保存路径: {relative_path}\n"
                result_text += f"📂 MCP服务目录: {service_dir}\n"
                result_text += f"📊 表名: {schema}.{table_name}\n"
                result_text += f"📝 导出类型: {export_type}\n"
                if export_type in ["data", "both"]:
                    result_text += f"📋 数据行数: {len(data_results) if 'data_results' in locals() else 0}\n"
                result_text += f"⏰ 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                
                return [TextContent(type="text", text=result_text)]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"导出Excel时发生错误: {str(e)}"
                )]
        
        elif name == "get_cache_info":
            try:
                cache_info = db.get_cache_info()
                
                result_text = "📊 查询缓存统计信息:\n\n"
                result_text += f"缓存大小: {cache_info['cache_size']} / {cache_info['max_size']} 条\n"
                result_text += f"缓存TTL: {cache_info['ttl']} 秒\n"
                result_text += f"缓存条目数: {len(cache_info['entries'])} 个\n\n"
                
                if cache_info['entries']:
                    result_text += "📋 缓存条目:\n"
                    for i, entry in enumerate(cache_info['entries'][:10], 1):  # 只显示前10个
                        result_text += f"{i}. {entry[:20]}...\n"
                    if len(cache_info['entries']) > 10:
                        result_text += f"... 还有 {len(cache_info['entries']) - 10} 个条目\n"
                
                return [TextContent(type="text", text=result_text)]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"获取缓存信息失败: {str(e)}"
                )]
        
        elif name == "clear_cache":
            try:
                db.clear_cache()
                return [TextContent(
                    type="text",
                    text="✅ 查询缓存已清空"
                )]
                
            except Exception as e:
                return [TextContent(
                    type="text",
                    text=f"清空缓存失败: {str(e)}"
                )]
        
        else:
            return create_error_response(f"未知的工具: {name}", "工具错误")
    
    except Exception as e:
        return create_error_response(f"工具调用失败: {str(e)}", "系统错误")

async def main():
    """主函数"""
    # 初始化配置和数据库连接测试
    logger.info("启动达梦数据库MCP服务...")
    
    try:
        # 获取配置信息
        config = get_config_instance()
        logger.info(f"配置加载成功，安全模式: {config.security_mode.value}")
        
        # 获取数据库实例并测试连接
        db = get_db_instance()
        if db.test_connection():
            logger.info("达梦数据库连接测试成功")
        else:
            logger.warning("达梦数据库连接测试失败，服务仍将启动")
            
    except Exception as e:
        logger.error(f"服务初始化失败: {e}")
        logger.error("请检查Cursor MCP配置中的环境变量设置")
        sys.exit(1)
    
    # 检查是否在HTTP模式下运行（Smithery.ai平台）
    if os.getenv("MCP_SERVER_MODE") == "http":
        # HTTP模式 - 用于Smithery.ai平台
        logger.info("启动HTTP模式MCP服务...")
        fastmcp = FastMCP("dm-mcp")
        
        # 注册所有工具
        @fastmcp.tool()
        async def test_connection() -> str:
            """测试达梦数据库连接"""
            result = db.test_connection()
            return f"达梦数据库连接测试: {'成功' if result else '失败'}"
        
        @fastmcp.tool()
        async def get_security_info() -> str:
            """获取当前安全配置信息"""
            security_info = db.get_security_info()
            info_text = "当前安全配置信息:\n\n"
            info_text += f"安全模式: {security_info['security_mode']}\n"
            info_text += f"只读模式: {'是' if security_info['readonly_mode'] else '否'}\n"
            info_text += f"允许写入操作: {'是' if security_info['write_allowed'] else '否'}\n"
            info_text += f"允许危险操作: {'是' if security_info['dangerous_operations_allowed'] else '否'}\n"
            info_text += f"允许访问的模式: {', '.join(security_info['allowed_schemas'])}\n"
            info_text += f"最大返回行数: {security_info['max_result_rows']}\n"
            info_text += f"查询日志: {'启用' if security_info['query_log_enabled'] else '禁用'}\n"
            return info_text
        
        # 启动HTTP服务器
        import uvicorn
        uvicorn.run(fastmcp, host="0.0.0.0", port=8000)
    else:
        # 标准stdio模式 - 用于本地Cursor
        logger.info("启动stdio模式MCP服务...")
        async with stdio_server() as (read_stream, write_stream):
            await server.run(
                read_stream,
                write_stream,
                InitializationOptions(
                    server_name="dm-mcp",
                    server_version="1.0.0",
                    capabilities=server.get_capabilities(
                        notification_options=NotificationOptions(),
                        experimental_capabilities={}
                    )
                )
            )

if __name__ == "__main__":
    asyncio.run(main()) 